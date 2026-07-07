#!/usr/bin/env python3
"""
Oppo A33w Sensor Data Exporter & Converter
Converts raw multi-sensor CSV records into modern multi-sheet Excel (.xlsx) workbooks
separated by sensor type, with summary statistics and formatted columns.
"""

import sys
import os
import argparse
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is required. Run 'pip install openpyxl'")
    sys.exit(1)

def convert_csv_to_excel(csv_path, xlsx_path):
    if not os.path.exists(csv_path):
        print(f"Error: CSV file '{csv_path}' not found.")
        return

    print(f"Converting '{csv_path}' -> '{xlsx_path}'...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Dictionary to group rows by Sensor Type
    sensor_groups = {}
    sensor_names = {
        1: "Accelerometer",
        2: "Magnetometer",
        4: "Gyroscope",
        5: "Light Sensor",
        8: "Proximity Sensor"
    }

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for row in reader:
            if len(row) >= 4:
                try:
                    stype = int(row[1])
                    if stype not in sensor_groups:
                        sensor_groups[stype] = []
                    sensor_groups[stype].append(row)
                except ValueError:
                    pass

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for stype, rows in sensor_groups.items():
        sheet_name = sensor_names.get(stype, f"Sensor_Type_{stype}")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Title Block
        ws.append([f"Oppo A33w Research Dataset: {sheet_name}"])
        ws['A1'].font = Font(size=14, bold=True, color="1E3A8A")
        ws.append([]) # Empty row

        # Columns
        headers = ["Timestamp (ms)", "Relative Time (s)", "Sensor Hardware Name", "Value X", "Value Y", "Value Z"]
        ws.append(headers)

        # Style header row (Row 3)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        start_ts = float(rows[0][0]) if rows else 0.0

        for r in rows:
            ts = float(r[0])
            rel_sec = (ts - start_ts) / 1000.0
            name = r[2].strip('"')
            vx = float(r[3])
            vy = float(r[4]) if len(r) > 4 else 0.0
            vz = float(r[5]) if len(r) > 5 else 0.0

            ws.append([ts, round(rel_sec, 4), name, vx, vy, vz])

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(xlsx_path)
    print(f"Successfully generated multi-sheet Excel workbook: {xlsx_path}")

def main():
    parser = argparse.ArgumentParser(description="Convert Oppo A33w CSV Data to Excel (.xlsx)")
    parser.add_argument("csv_file", help="Input CSV recording file")
    parser.add_argument("output_xlsx", nargs="?", help="Output Excel .xlsx file path")
    args = parser.parse_args()

    out_file = args.output_xlsx
    if not out_file:
        out_file = os.path.splitext(args.csv_file)[0] + "_processed.xlsx"

    convert_csv_to_excel(args.csv_file, out_file)

if __name__ == "__main__":
    main()
