#!/usr/bin/env python3
"""
Oppo A33w Batch Dataset Processor & Comparator
Ingests a directory of CSV experimental recordings, runs activity classification,
extracts summary statistics, and outputs a multi-sheet Master Comparison Excel (.xlsx) workbook.
"""

import sys
import os
import glob
import argparse
import csv

try:
    import numpy as np
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: numpy and openpyxl are required.")
    sys.exit(1)

def process_batch(directory_path, output_xlsx="master_batch_comparison.xlsx"):
    if not os.path.exists(directory_path):
        print(f"Error: Directory '{directory_path}' does not exist.")
        return

    csv_files = glob.glob(os.path.join(directory_path, "*.csv"))
    if not csv_files:
        print(f"No CSV files found in '{directory_path}'.")
        return

    print(f"Batch processing {len(csv_files)} CSV datasets from '{directory_path}' -> '{output_xlsx}'...")

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Batch Trial Comparison"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="38BDF8", bold=True)

    ws_summary.append(["Oppo A33w Master Trial Comparison Report"])
    ws_summary['A1'].font = Font(size=16, bold=True, color="0F172A")
    ws_summary.append([])

    headers = [
        "File Name", "Total Packets", "Duration (s)", "Effective Rate (Hz)",
        "Acc X RMS", "Acc Y RMS", "Acc Z RMS",
        "Dynamic RMS Vibration", "Peak-to-Peak (m/s²)",
        "Detected Steps", "Estimated Cadence (SPM)", "Classified Activity State"
    ]
    ws_summary.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 4
    for file_path in sorted(csv_files):
        timestamps = []
        acc_x, acc_y, acc_z = [], [], []

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if len(row) >= 6:
                    try:
                        stype = int(row[1])
                        if stype == 1:
                            timestamps.append(float(row[0]))
                            acc_x.append(float(row[3]))
                            acc_y.append(float(row[4]))
                            acc_z.append(float(row[5]))
                    except ValueError:
                        pass

        if not timestamps:
            continue

        ts_arr = np.array(timestamps)
        xa = np.array(acc_x)
        ya = np.array(acc_y)
        za = np.array(acc_z)

        duration = (ts_arr[-1] - ts_arr[0]) / 1000.0 if len(ts_arr)>1 else 0.0
        rate = len(ts_arr) / duration if duration > 0 else 0.0

        rms_x = np.sqrt(np.mean(xa**2))
        rms_y = np.sqrt(np.mean(ya**2))
        rms_z = np.sqrt(np.mean(za**2))

        # Combined magnitude
        mags = np.sqrt(xa**2 + ya**2 + za**2)
        dyn_acc = mags - np.mean(mags)
        dyn_rms = np.sqrt(np.mean(dyn_acc**2))
        p2p = np.max(mags) - np.min(mags)

        # Simple step counting
        std_acc = np.std(dyn_acc)
        threshold = max(0.4, 0.8 * std_acc)
        steps = 0
        last_step_time = 0
        for i in range(1, len(dyn_acc) - 1):
            if dyn_acc[i] > threshold and dyn_acc[i] > dyn_acc[i-1] and dyn_acc[i] > dyn_acc[i+1]:
                if (ts_arr[i] - last_step_time) >= 250.0:
                    steps += 1
                    last_step_time = ts_arr[i]

        cadence = (steps / duration) * 60.0 if duration > 0 else 0.0

        if dyn_rms < 0.15: state = "STATIONARY"
        elif dyn_rms < 1.2: state = "WALKING"
        elif dyn_rms < 3.5: state = "RUNNING / VIGOROUS"
        else: state = "HIGH VIBRATION"

        ws_summary.append([
            os.path.basename(file_path), len(ts_arr), round(duration, 2), round(rate, 1),
            round(rms_x, 4), round(rms_y, 4), round(rms_z, 4),
            round(dyn_rms, 4), round(p2p, 4), steps, round(cadence, 1), state
        ])
        row_idx += 1

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 13)

    wb.save(output_xlsx)
    print(f"Successfully generated Master Comparison Excel workbook: {output_xlsx}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Oppo A33w Batch Dataset Processor")
    parser.add_argument("directory", help="Directory containing CSV experimental trials")
    parser.add_argument("--output", default="master_batch_comparison.xlsx", help="Output Excel workbook path")
    args = parser.parse_args()
    process_batch(args.directory, args.output)
